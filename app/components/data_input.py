from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import streamlit as st

from openpyxl import load_workbook


class InputBase:
    sep: str
    user_input: Any

    def __init__(self, key=None):
        self.key = f"{self.__class__.__name__}-{key}"

    def parse(self):
        pass

    def seperator(self):
        sep_options = {
            "Tab (\\t)": "\t",
            "Comma (,)": ",",
            "Space (' ')": " "
        }
        user_sep = st.selectbox("Seperator", key=self.key,
                                options=sep_options.keys())
        self.sep = sep_options[user_sep]


@st.cache_data
def parse_file(file, header=False, index=False,
               sheet_name=0):

    index_col = None if not index else 0
    suffix = file.name.split(".")[-1]
    if suffix in ["csv", "txt", "tsv"]:
        header = None if not header else "infer"
        reader = pd.read_csv
        sep = "," if suffix == "csv" else "\t"
        kws = dict(sep=sep, header=header, index_col=index_col)
    else:
        header = None if not header else 0
        reader = pd.read_excel
        kws = dict(header=header, index_col=index_col, sheet_name=sheet_name)
    data = reader(file, **kws)
    return data


class FileUpload(InputBase):

    def __init__(self, key=None, header=False, index=False,
                 use_header=False, use_index=False):
        super().__init__(key=key)
        self.header = use_header
        self.index = use_index
        self.user_input = st.file_uploader("Choose a table file",
                                           key=f"table_reader-{self.key}",
                                           accept_multiple_files=False,
                                           label_visibility="collapsed",
                                           type=["txt", "csv", "xlsx"])

        if index & header:
            h1, h2 = st.columns(2)
            with h1:
                self.header = self._header_checkbox()
            with h2:
                self.index = self._index_checkbox()

        elif header:
            self.header = self._header_checkbox()

        elif index:
            self.index = self._index_checkbox()

        self.sheet_name = 0
        if self.user_input is not None:
            suffix = self.user_input.name.split(".")[-1]
            if suffix == "xlsx":
                wb = load_workbook(
                    self.user_input, read_only=True, keep_links=False)
                sheetnames = wb.sheetnames
                if len(sheetnames) > 1:
                    self.sheet_name = \
                        st.selectbox("Please select a sheet",
                                     options=sheetnames)

    def _header_checkbox(self):
        return st.checkbox("Use first row as header?",
                           value=self.header,
                           key=f"select-header-{self.key}")

    def _index_checkbox(self):
        return st.checkbox("Use first column as row labels?",
                           value=self.index,
                           key=f"select-index-{self.key}")

    def _parse_to_df(self):
        if self.user_input is not None:
            return parse_file(self.user_input, header=self.header,
                              index=self.index, sheet_name=self.sheet_name)

    def parse(self) -> np.ndarray:
        if self.user_input is not None:
            data = self._parse_to_df()
            if len(data.columns) == 1:
                return data.to_numpy().flatten()
            return data.to_numpy()

    def parse_parts(self, row_label=True, col_label=True):
        if self.user_input is not None:
            data = parse_file(self.user_input, header=col_label,
                              index=row_label, sheet_name=self.sheet_name)
            row = data.index.to_numpy(dtype=str)
            col = data.columns.to_numpy(dtype=str)
            data = data.to_numpy()
            return data, row, col

    def parse_dataframe(self) -> pd.DataFrame:
        return self._parse_to_df()

    @property
    def name(self):
        if self.user_input is None:
            return ""
        else:
            return Path(self.user_input.name).stem


class PasteText(InputBase):

    def __init__(self, cast_number=True, key=None):
        super().__init__(key=key)
        self.cast_number = cast_number
        self.user_input = st.text_area("Paste data here", key=self.key)
        self.seperator()

    def parse(self):
        if len(self.user_input) > 0:
            rows = self.user_input.strip().split("\n")
            raw_data = [row.split(self.sep) for row in rows]
            try:
                data = np.array(raw_data)
                if self.cast_number:
                    data = data.astype(float)
                return pd.DataFrame(data)
            except Exception:
                st.error("Your seperator seems incorrect")
